"""
Data Bot — Populates "Faults vs Tons 2026.xlsx" from alarm and production CSVs.

Spreadsheet structure (per machine sheet e.g. CM08):
  Row 1:  blank
  Row 2:  Machine name
  Row 3:  Crew headers  (B Crew col B, D Crew col S, C Crew col AJ, A Crew col BA)
  Row 4:  Column headers (Shift, Date, M1443..M1461, Total, Tons Cut, M2255 x4)
  Row 5+: Data rows (one row per date)

Each crew section is 17 columns wide:
  offset 0:  Shift (DS or NS)
  offset 1:  Date
  offsets 2-13: Fault counts for motors M1443,M1447,M1450,M1454,M1458,M1463,
                                          M1441,M1445,M1448,M1452,M1456,M1461
  offset 14: Total trips (=SUM formula)
  offset 15: Tons Cut
  offset 16: M2255 Faults

NAME YOUR CSV FILES LIKE THIS:
  CM08_B_alarms.csv       ← alarm data for CM08, B Crew
  CM08_D_alarms.csv       ← alarm data for CM08, D Crew
  CM08_B_production.csv   ← production (tonnes) for CM08, B Crew
  CM03_A_alarms.csv       ← alarm data for CM03, A Crew

The machine (CM02/CM03/CM05/CM07/CM08/CM09) and crew (A/B/C/D) are read
from the filename automatically.

ALARM CSV columns required:
  STATE, AE_TIMESTAMP, EVENT_CODE, EVENT_DESCRIPTION, SEQUENCE_NO

PRODUCTION CSV columns required:
  DATE (or TIMESTAMP), SHIFT (DS/NS), TONNES (or MTRS for metres)
  Optional: MACHINE_ID, CREW

Usage:
  python main.py --input ./input_data --output "Faults vs Tons 2026.xlsx"
"""

import argparse
import os
import re
import glob
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from datetime import datetime, date

# ---------------------------------------------------------------------------
# Spreadsheet layout constants
# ---------------------------------------------------------------------------

# Valid machine sheet names in the workbook
MACHINE_SHEETS = ["CM02", "CM03", "CM05", "CM07", "CM08", "CM09"]

# Crew → starting column (1-based) in the machine sheets
CREW_START_COL = {"B": 2, "D": 19, "C": 36, "A": 53}

# Column offsets within a crew section (0-based from crew start col)
COL_SHIFT   = 0
COL_DATE    = 1
COL_FAULTS_START = 2   # first fault motor column
COL_FAULTS_END   = 13  # last fault motor column (inclusive, 12 motors total)
COL_TOTAL   = 14
COL_TONS    = 15
COL_M2255   = 16
SECTION_WIDTH = 17

# Motor code → fault column offset within fault block (offset 2..13)
# Order matches the column headers in the spreadsheet
MOTOR_OFFSET = {
    "M1443": 2,   # Jam Overload, Left Cutter
    "M1447": 3,   # Jam Overload, Right Cutter
    "M1450": 4,   # Jam Overload, Left Gath Head
    "M1454": 5,   # Jam Overload, Right Gath Head
    "M1458": 6,   # Jam Overload, Left Traction
    "M1463": 7,   # Jam Overload, Right Traction
    "M1441": 8,   # Thermal Overload, Left Cutter
    "M1445": 9,   # Thermal Overload, RH Cutter
    "M1448": 10,  # Thermal Overload, LH Gath Head
    "M1452": 11,  # Thermal Overload, RH Gath Head
    "M1456": 12,  # Thermal Overload, Left Traction
    "M1461": 13,  # Thermal Overload, Right Traction
}

# Data rows start at row 5 in machine sheets
DATA_START_ROW = 5

# Day shift hours (06:00 inclusive to 18:00 exclusive = DS, else NS)
DS_START_HOUR = 6
DS_END_HOUR   = 18

# Metres to tonnes conversion (from spreadsheet formula: MTRS * 11.9 * 2)
METRES_TO_TONNES = 11.9 * 2


# ---------------------------------------------------------------------------
# Filename parsing
# ---------------------------------------------------------------------------

def parse_filename(path: str):
    """
    Extract machine ID, crew, and CSV type from filename.
    Returns (machine, crew, csv_type) where:
      machine  = 'CM08' etc, or None
      crew     = 'A'/'B'/'C'/'D', or None
      csv_type = 'alarm' or 'production'
    """
    name = os.path.splitext(os.path.basename(path))[0].upper()

    machine = None
    for m in MACHINE_SHEETS:
        if m in name:
            machine = m
            break

    crew = None
    # Look for _B_ / _D_ / _C_ / _A_ pattern
    crew_match = re.search(r'[_\-\s]([ABCD])[_\-\s]', name)
    if crew_match:
        crew = crew_match.group(1)

    csv_type = "production" if "PROD" in name or "TONNE" in name or "TONS" in name else "alarm"

    return machine, crew, csv_type


# ---------------------------------------------------------------------------
# Shift detection from timestamp
# ---------------------------------------------------------------------------

def get_shift(ts) -> str:
    """Return 'DS' or 'NS' based on hour of timestamp."""
    if hasattr(ts, "hour"):
        return "DS" if DS_START_HOUR <= ts.hour < DS_END_HOUR else "NS"
    return "DS"


def to_date(ts) -> date:
    """Extract date portion, adjusting NS times past midnight back one day."""
    if hasattr(ts, "hour"):
        # Night shift hours 00:00-05:59 belong to the previous calendar day
        if ts.hour < DS_START_HOUR:
            return (ts - pd.Timedelta(days=1)).date()
        return ts.date()
    return ts


# ---------------------------------------------------------------------------
# Workbook helpers
# ---------------------------------------------------------------------------

def open_workbook(path: str):
    if not os.path.exists(path):
        raise FileNotFoundError(
            f"Workbook not found: {path}\n"
            "Place 'Faults vs Tons 2026.xlsx' in the data_bot folder or use --output to specify the path."
        )
    print(f"Opening workbook: {path}")
    return load_workbook(path)


def find_date_row(ws, target_date: date) -> int:
    """
    Search col C (B Crew date column) for target_date.
    Returns row number if found, else 0.
    """
    for row in ws.iter_rows(min_row=DATA_START_ROW, min_col=3, max_col=3, values_only=False):
        cell = row[0]
        val = cell.value
        if val is None:
            continue
        cell_date = val.date() if hasattr(val, "date") else val
        if cell_date == target_date:
            return cell.row
    return 0


def add_date_row(ws, target_date: date, crew: str, shift: str) -> int:
    """
    Append a new data row for the given date and return its row number.
    Writes Shift + Date for all crew sections and inserts Total formulas.
    """
    # Find next empty row (first row where col C is blank from DATA_START_ROW)
    new_row = DATA_START_ROW
    for row in ws.iter_rows(min_row=DATA_START_ROW, min_col=3, max_col=3):
        if row[0].value is None:
            new_row = row[0].row
            break
    else:
        new_row = ws.max_row + 1

    # Write Shift and Date for every crew section
    for c, start_col in CREW_START_COL.items():
        ws.cell(new_row, start_col + COL_SHIFT).value = shift
        ws.cell(new_row, start_col + COL_DATE).value  = target_date
        # Total formula
        fault_start_col = get_column_letter(start_col + COL_FAULTS_START)
        fault_end_col   = get_column_letter(start_col + COL_FAULTS_END)
        ws.cell(new_row, start_col + COL_TOTAL).value = \
            f"=SUM({fault_start_col}{new_row}:{fault_end_col}{new_row})"
        # Zero out fault cols
        for offset in range(COL_FAULTS_START, COL_FAULTS_END + 1):
            if ws.cell(new_row, start_col + offset).value is None:
                ws.cell(new_row, start_col + offset).value = 0

    return new_row


# ---------------------------------------------------------------------------
# Crew lookup — find which crew is rostered for a date+shift from the sheet
# ---------------------------------------------------------------------------

def find_crew_for_date_shift(ws, target_date: date, shift: str) -> str | None:
    """
    The spreadsheet has the crew rotation pre-populated.
    For a given date + shift, scan every crew section's Shift+Date columns
    and return the crew letter whose section matches.
    Returns None if no matching crew section is found.
    """
    for crew, start_col in CREW_START_COL.items():
        for row_num in range(DATA_START_ROW, ws.max_row + 1):
            date_val  = ws.cell(row_num, start_col + COL_DATE).value
            shift_val = ws.cell(row_num, start_col + COL_SHIFT).value
            if date_val is None:
                continue
            cell_date = date_val.date() if hasattr(date_val, "date") else date_val
            if cell_date == target_date and str(shift_val).strip().upper() == shift.upper():
                return crew
    return None


# ---------------------------------------------------------------------------
# Write alarm counts to machine sheet
# ---------------------------------------------------------------------------

def write_alarms_to_sheet(ws, counts: dict, crew: str):
    """
    counts: {(date, shift): {motor_code: count}}
    crew:   'A'/'B'/'C'/'D'
    """
    start_col = CREW_START_COL.get(crew)
    if start_col is None:
        print(f"  WARNING: Unknown crew '{crew}', skipping.")
        return

    written = 0
    for (row_date, shift), motor_counts in sorted(counts.items()):
        row_num = find_date_row(ws, row_date)
        if row_num == 0:
            row_num = add_date_row(ws, row_date, crew, shift)

        # Update shift for this crew
        ws.cell(row_num, start_col + COL_SHIFT).value = shift

        for motor_code, count in motor_counts.items():
            offset = MOTOR_OFFSET.get(motor_code.upper())
            if offset is None:
                continue
            col = start_col + offset
            existing = ws.cell(row_num, col).value
            existing = int(existing) if isinstance(existing, (int, float, str)) and str(existing).strip().lstrip('-').isdigit() else 0
            ws.cell(row_num, col).value = existing + count
            written += count

    print(f"  Wrote {written} fault counts to {ws.title} — {crew} Crew.")


# ---------------------------------------------------------------------------
# Process alarm CSV
# ---------------------------------------------------------------------------

def process_alarm_csv(df: pd.DataFrame, wb, machine: str, crew: str):
    if machine not in wb.sheetnames:
        print(f"  WARNING: Sheet '{machine}' not found in workbook. Skipping.")
        return

    df.columns = [c.upper().strip() for c in df.columns]
    required = {"STATE", "AE_TIMESTAMP", "EVENT_CODE"}
    if not required.issubset(set(df.columns)):
        print(f"  WARNING: Missing required columns {required - set(df.columns)}. Skipping.")
        return

    # Parse timestamps
    df["AE_TIMESTAMP"] = pd.to_datetime(df["AE_TIMESTAMP"], dayfirst=True, errors="coerce")
    df = df.dropna(subset=["AE_TIMESTAMP"])

    # Only count "On" events (each On = one trip/fault occurrence)
    df = df[df["STATE"].astype(str).str.strip() == "1"]

    if df.empty:
        print("  No 'On' alarm events found.")
        return

    # Aggregate: {(date, shift): {motor_code: count}}
    counts = {}
    for _, row in df.iterrows():
        ts   = row["AE_TIMESTAMP"]
        d    = to_date(ts)
        sh   = get_shift(ts)
        code = str(row["EVENT_CODE"]).upper().strip()
        key  = (d, sh)
        counts.setdefault(key, {})
        counts[key][code] = counts[key].get(code, 0) + 1

    ws = wb[machine]
    write_alarms_to_sheet(ws, counts, crew)


# ---------------------------------------------------------------------------
# Process production CSV
# ---------------------------------------------------------------------------

def process_production_csv(df: pd.DataFrame, wb, machine: str, crew: str = None):
    """
    crew is optional. If not supplied (or None), the bot looks up which crew
    is rostered for each date+shift from the pre-populated spreadsheet rotation.
    """
    if machine not in wb.sheetnames:
        print(f"  WARNING: Sheet '{machine}' not found. Skipping production data.")
        return

    df.columns = [c.upper().strip() for c in df.columns]

    # Find timestamp/date column
    ts_col = next((c for c in df.columns if "TIME" in c or "DATE" in c), None)
    if ts_col is None:
        print("  WARNING: No date/timestamp column found in production CSV. Skipping.")
        return

    df[ts_col] = pd.to_datetime(df[ts_col], dayfirst=True, errors="coerce")
    df = df.dropna(subset=[ts_col])

    # Find tonnes column
    tons_col = next((c for c in df.columns if "TONNE" in c or "TONS" in c), None)
    mtrs_col = next((c for c in df.columns if "MTR" in c or "METER" in c), None)

    if tons_col:
        df["_TONS"] = pd.to_numeric(df[tons_col], errors="coerce").fillna(0)
    elif mtrs_col:
        df["_TONS"] = pd.to_numeric(df[mtrs_col], errors="coerce").fillna(0) * METRES_TO_TONNES
        print(f"  Converting metres to tonnes (×{METRES_TO_TONNES})")
    else:
        print("  WARNING: No TONNES or MTRS column found. Skipping production data.")
        return

    # Shift column
    shift_col = next((c for c in df.columns if "SHIFT" in c), None)

    ws = wb[machine]
    written = 0
    skipped = 0
    for _, row in df.iterrows():
        ts    = row[ts_col]
        d     = to_date(ts)
        shift = row[shift_col].strip().upper() if shift_col and pd.notna(row[shift_col]) else get_shift(ts)
        tons  = row["_TONS"]

        # Determine crew: use supplied crew, or look it up from the spreadsheet rotation
        resolved_crew = crew
        if resolved_crew is None:
            resolved_crew = find_crew_for_date_shift(ws, d, shift)
        if resolved_crew is None:
            print(f"  WARNING: No crew found in spreadsheet for {d} {shift} — skipping row.")
            skipped += 1
            continue

        start_col = CREW_START_COL[resolved_crew]
        row_num = find_date_row(ws, d)
        if row_num == 0:
            row_num = add_date_row(ws, d, resolved_crew, shift)

        ws.cell(row_num, start_col + COL_SHIFT).value = shift
        ws.cell(row_num, start_col + COL_TONS).value  = round(float(tons), 2)
        written += 1

    msg = f"  Wrote {written} production rows to {machine}"
    if crew:
        msg += f" — {crew} Crew."
    else:
        msg += " (crew auto-detected from spreadsheet rotation)."
    if skipped:
        msg += f" {skipped} rows skipped (no matching crew in rotation)."
    print(msg)


# ---------------------------------------------------------------------------
# CSV type detection
# ---------------------------------------------------------------------------

def detect_csv_type(df: pd.DataFrame) -> str:
    cols = {c.upper().strip() for c in df.columns}
    if {"STATE", "AE_TIMESTAMP", "EVENT_CODE"}.issubset(cols):
        return "alarm"
    if {"TONNES", "MTRS"}.intersection(cols) or {"TONNES", "METERS"}.intersection(cols):
        return "production"
    return "unknown"


# ---------------------------------------------------------------------------
# Zero-ton cleanup — erase faults on rows with 0 tons cut
# ---------------------------------------------------------------------------

def clear_zero_ton_faults(wb):
    """
    For every data row in every machine sheet:
    if a crew section has Tons Cut = 0 or blank, zero out all 12 fault columns
    for that crew. This prevents phantom faults on non-operating shifts.
    """
    cleared = 0
    for sheet_name in MACHINE_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for row_num in range(DATA_START_ROW, ws.max_row + 1):
            for crew, start_col in CREW_START_COL.items():
                tons_cell = ws.cell(row_num, start_col + COL_TONS)
                tons_val  = tons_cell.value
                # Treat 0, None, blank string, or non-numeric as zero-tons
                try:
                    tons = float(tons_val) if tons_val not in (None, "", " ") else 0.0
                except (ValueError, TypeError):
                    tons = 0.0

                if tons == 0.0:
                    # Zero out all 12 fault columns
                    for offset in range(COL_FAULTS_START, COL_FAULTS_END + 1):
                        cell = ws.cell(row_num, start_col + offset)
                        if cell.value not in (None, 0, ""):
                            cell.value = 0
                            cleared += 1
                    # Also zero out the Tons Cut cell itself
                    if tons_cell.value not in (None, 0, ""):
                        tons_cell.value = 0
                        cleared += 1

    if cleared:
        print(f"Zero-ton cleanup: cleared {cleared} fault cell(s) on rows with 0 tons cut.")
    else:
        print("Zero-ton cleanup: no faults needed clearing (all rows with faults had tons).")


# ---------------------------------------------------------------------------
# Fill missing machines — write zero rows for dates not covered by a machine
# ---------------------------------------------------------------------------

def fill_missing_machines(wb):
    """
    After all CSVs are processed, scan every machine sheet to build a master
    list of (date, shift) pairs. Then, for any machine sheet that is missing
    one of those dates entirely, insert a zero row so every machine has a
    record for every date in the dataset.
    """
    # Step 1 — collect every (date, shift) seen across all machine sheets
    # Use a dict so we keep the shift for each date (DS preferred over NS)
    all_date_shifts = {}  # date -> shift
    for sheet_name in MACHINE_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for row_num in range(DATA_START_ROW, ws.max_row + 1):
            date_val  = ws.cell(row_num, CREW_START_COL["B"] + COL_DATE).value
            shift_val = ws.cell(row_num, CREW_START_COL["B"] + COL_SHIFT).value
            if date_val is None:
                continue
            d = date_val.date() if hasattr(date_val, "date") else date_val
            # Prefer DS if we've already seen a DS for this date
            if d not in all_date_shifts or shift_val == "DS":
                all_date_shifts[d] = shift_val or "DS"

    if not all_date_shifts:
        return

    # Step 2 — for each machine sheet, find missing dates and insert zero rows
    filled = 0
    for sheet_name in MACHINE_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]

        existing_dates = set()
        for row_num in range(DATA_START_ROW, ws.max_row + 1):
            date_val = ws.cell(row_num, CREW_START_COL["B"] + COL_DATE).value
            if date_val is None:
                continue
            d = date_val.date() if hasattr(date_val, "date") else date_val
            existing_dates.add(d)

        missing = sorted(d for d in all_date_shifts if d not in existing_dates)
        for d in missing:
            shift = all_date_shifts[d]
            add_date_row(ws, d, "B", shift)  # writes Shift+Date+formulas for all 4 crews
            # Explicitly zero tons for all crew sections on this row
            row_num = find_date_row(ws, d)
            if row_num:
                for start_col in CREW_START_COL.values():
                    ws.cell(row_num, start_col + COL_TONS).value = 0
            filled += 1

    if filled:
        print(f"Missing machines: filled {filled} date row(s) with zeros across machines with no data.")
    else:
        print("Missing machines: all machines already have rows for every date.")


# ---------------------------------------------------------------------------
# Core run function (callable from CLI and web server)
# ---------------------------------------------------------------------------

def run(input_dir: str, output_path: str):
    """Process all CSVs in input_dir and write results to output_path."""
    csv_files = glob.glob(os.path.join(input_dir, "*.csv"))
    if not csv_files:
        print(f"No CSV files found in: {input_dir}")
        print("Drop your alarm and production CSV files into that folder and run again.")
        return

    print(f"Found {len(csv_files)} CSV file(s)\n")

    wb = open_workbook(output_path)   # raises FileNotFoundError if missing

    for csv_path in sorted(csv_files):
        fname = os.path.basename(csv_path)
        print(f"Processing: {fname}")

        machine, crew, file_type = parse_filename(csv_path)

        if machine is None:
            print(f"  SKIPPED — could not detect machine name (CM02/CM03/etc.) in filename.")
            print()
            continue
        if crew is None:
            print(f"  SKIPPED — could not detect crew (A/B/C/D) in filename. "
                  f"Rename to e.g. '{machine}_B_alarms.csv'")
            print()
            continue

        print(f"  Machine={machine}  Crew={crew}  Type={file_type}")

        try:
            df = pd.read_csv(csv_path)
        except Exception as e:
            print(f"  ERROR reading CSV: {e}")
            print()
            continue

        if file_type == "alarm":
            process_alarm_csv(df, wb, machine, crew)
        else:
            process_production_csv(df, wb, machine, crew)

        print()

    fill_missing_machines(wb)
    clear_zero_ton_faults(wb)
    wb.save(output_path)
    print(f"✓ Saved: {output_path}")


# ---------------------------------------------------------------------------
# Main (CLI entry point)
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Populate 'Faults vs Tons 2026.xlsx' from alarm and production CSVs.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
FILE NAMING (machine and crew are read from the filename):
  CM08_B_alarms.csv        → CM08 sheet, B Crew, alarm data
  CM08_D_alarms.csv        → CM08 sheet, D Crew, alarm data
  CM03_A_production.csv    → CM03 sheet, A Crew, production data

SUPPORTED MACHINES: CM02, CM03, CM05, CM07, CM08, CM09
SUPPORTED CREWS:    A, B, C, D
        """
    )
    parser.add_argument("--input",  default="./input_data",
                        help="Folder containing CSV files (default: ./input_data)")
    parser.add_argument("--output", default="Faults vs Tons 2026.xlsx",
                        help="Path to the Excel workbook")
    args = parser.parse_args()

    try:
        run(args.input, args.output)
    except FileNotFoundError as e:
        print(f"ERROR: {e}")


if __name__ == "__main__":
    main()
