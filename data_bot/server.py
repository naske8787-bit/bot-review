"""
Data Bot Web Server
Run: python server.py
Then open http://localhost:5050 in your browser.
"""

import os
import sys
import io
import importlib
from contextlib import redirect_stdout

from flask import Flask, request, jsonify, send_from_directory

# Ensure main.py is importable from the same directory
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

BASE_DIR    = os.path.dirname(os.path.abspath(__file__))
INPUT_DIR   = os.path.join(BASE_DIR, "input_data")
STATIC_DIR  = os.path.join(BASE_DIR, "static")

# Default spreadsheet path — works on both Linux dev container and Windows
if sys.platform == "win32":
    DEFAULT_OUTPUT = os.path.join(
        os.environ.get("USERPROFILE", "C:\\Users\\Default"),
        "OneDrive - Komatsu Ltd", "Documents", "Miners", "Faults vs Tons 2026.xlsx"
    )
else:
    DEFAULT_OUTPUT = os.path.join(BASE_DIR, "Faults vs Tons 2026.xlsx")

os.makedirs(INPUT_DIR, exist_ok=True)

app = Flask(__name__, static_folder=STATIC_DIR)


# ---------------------------------------------------------------------------
# Static UI
# ---------------------------------------------------------------------------

@app.route("/")
def index():
    return send_from_directory(STATIC_DIR, "index.html")


# ---------------------------------------------------------------------------
# File management
# ---------------------------------------------------------------------------

@app.route("/files", methods=["GET"])
def list_files():
    files = sorted(f for f in os.listdir(INPUT_DIR) if f.lower().endswith(".csv"))
    return jsonify({"files": files})


@app.route("/upload", methods=["POST"])
def upload():
    uploaded = request.files.getlist("files")
    saved, errors = [], []
    for f in uploaded:
        name = os.path.basename(f.filename)  # strip any path
        if not name.lower().endswith(".csv"):
            errors.append(f"{name}: not a CSV — skipped")
            continue
        dest = os.path.join(INPUT_DIR, name)
        f.save(dest)
        saved.append(name)
    return jsonify({"saved": saved, "errors": errors})


@app.route("/files/<path:name>", methods=["DELETE"])
def delete_file(name):
    safe = os.path.basename(name)  # prevent path traversal
    path = os.path.join(INPUT_DIR, safe)
    if os.path.exists(path):
        os.remove(path)
        return jsonify({"ok": True})
    return jsonify({"ok": False, "error": "File not found"}), 404


@app.route("/clear", methods=["POST"])
def clear_files():
    removed = []
    for f in os.listdir(INPUT_DIR):
        if f.lower().endswith(".csv"):
            os.remove(os.path.join(INPUT_DIR, f))
            removed.append(f)
    return jsonify({"removed": removed})


# ---------------------------------------------------------------------------
# Process
# ---------------------------------------------------------------------------

@app.route("/process", methods=["POST"])
def process():
    data = request.get_json(silent=True) or {}
    output_path = (data.get("output_path") or DEFAULT_OUTPUT).strip()

    buf = io.StringIO()
    ok  = True

    try:
        import main as bot_main
        importlib.reload(bot_main)  # pick up any saved changes
        with redirect_stdout(buf):
            bot_main.run(INPUT_DIR, output_path)
    except FileNotFoundError as e:
        ok = False
        buf.write(f"\nERROR: {e}\n")
    except Exception as e:
        ok = False
        buf.write(f"\nUNEXPECTED ERROR: {e}\n")
        import traceback
        buf.write(traceback.format_exc())

    return jsonify({"ok": ok, "log": buf.getvalue()})


@app.route("/add_tons", methods=["POST"])
def add_tons():
    data        = request.get_json(silent=True) or {}
    output_path = (data.get("output_path") or DEFAULT_OUTPUT).strip()
    machine     = data.get("machine", "").strip().upper()
    date_str    = data.get("date", "").strip()
    shift       = data.get("shift", "DS").strip().upper()
    tons_raw    = data.get("tons", "")

    try:
        from datetime import datetime as dt
        import main as bot_main
        importlib.reload(bot_main)
        from openpyxl import load_workbook

        if machine not in bot_main.MACHINE_SHEETS:
            return jsonify({"ok": False, "log": f"Unknown machine: {machine}"}), 400

        d    = dt.strptime(date_str, "%Y-%m-%d").date()
        tons = float(str(tons_raw).strip())

        wb = load_workbook(output_path)
        ws = wb[machine]

        # Auto-detect crew from pre-populated spreadsheet rotation
        crew = bot_main.find_crew_for_date_shift(ws, d, shift)
        if crew is None:
            return jsonify({
                "ok": False,
                "log": f"No crew found in spreadsheet for {date_str} {shift}.\n"
                       f"Check that the crew rotation is pre-populated for this date in the {machine} sheet."
            }), 400

        row_num = bot_main.find_date_row(ws, d)
        if row_num == 0:
            row_num = bot_main.add_date_row(ws, d, crew, shift)

        start_col = bot_main.CREW_START_COL[crew]
        ws.cell(row_num, start_col + bot_main.COL_SHIFT).value = shift
        ws.cell(row_num, start_col + bot_main.COL_TONS).value  = round(tons, 2)

        # If 0 tonnes entered, wipe all fault columns for this row/crew too
        faults_cleared = 0
        if tons == 0.0:
            for offset in range(bot_main.COL_FAULTS_START, bot_main.COL_FAULTS_END + 1):
                cell = ws.cell(row_num, start_col + offset)
                if cell.value not in (None, 0, ""):
                    cell.value = 0
                    faults_cleared += 1

        wb.save(output_path)
        log = f"✓ Auto-detected crew: {crew} Crew\n" \
              f"  Wrote {tons:.2f} tonnes → {machine} / {crew} Crew / {date_str} / {shift}\n"
        if faults_cleared:
            log += f"  Cleared {faults_cleared} fault cell(s) (0 tonnes entered).\n"
        log += f"  Saved: {output_path}\n"
        return jsonify({"ok": True, "log": log})

    except ValueError as e:
        return jsonify({"ok": False, "log": f"ERROR: {e}"}), 400
    except FileNotFoundError as e:
        return jsonify({"ok": False, "log": f"ERROR: {e}"}), 400
    except Exception as e:
        import traceback
        return jsonify({"ok": False, "log": f"UNEXPECTED ERROR: {e}\n{traceback.format_exc()}"}), 500


# ---------------------------------------------------------------------------
# Generate report spreadsheet
# ---------------------------------------------------------------------------

@app.route("/generate_report", methods=["POST"])
def generate_report_endpoint():
    data        = request.get_json(silent=True) or {}
    source_path = (data.get("source_path") or DEFAULT_OUTPUT).strip()
    # Default report output sits next to the source file
    default_out = os.path.join(os.path.dirname(source_path),
                               "Faults vs Tons Report.xlsx")
    output_path = (data.get("output_path") or default_out).strip()

    try:
        import generate_report as gr
        importlib.reload(gr)
        log = gr.generate_report(source_path, output_path)
        return jsonify({"ok": True, "log": log, "output_path": output_path})
    except FileNotFoundError as e:
        return jsonify({"ok": False, "log": f"ERROR: {e}"}), 400
    except Exception as e:
        import traceback
        return jsonify({"ok": False, "log": f"UNEXPECTED ERROR: {e}\n{traceback.format_exc()}"}), 500


@app.route("/default_output")
def default_output():
    return jsonify({"path": DEFAULT_OUTPUT})


# ---------------------------------------------------------------------------
# Fill zeros for a date range
# ---------------------------------------------------------------------------

@app.route("/fill_zeros", methods=["POST"])
def fill_zeros():
    from datetime import date, timedelta
    data = request.get_json(silent=True) or {}
    output_path = (data.get("output_path") or DEFAULT_OUTPUT).strip()
    date_from   = data.get("date_from", "").strip()
    date_to     = data.get("date_to",   "").strip()
    machine_sel = data.get("machine", "ALL").strip().upper()  # "ALL" or e.g. "CM08"
    buf = io.StringIO()
    ok  = True

    try:
        from datetime import datetime as dt
        d_from = dt.strptime(date_from, "%Y-%m-%d").date()
        d_to   = dt.strptime(date_to,   "%Y-%m-%d").date()
        if d_from > d_to:
            d_from, d_to = d_to, d_from

        import main as bot_main
        importlib.reload(bot_main)

        from openpyxl import load_workbook
        wb = load_workbook(output_path)

        all_dates = [d_from + timedelta(days=i) for i in range((d_to - d_from).days + 1)]
        filled = 0

        target_sheets = bot_main.MACHINE_SHEETS if machine_sel == "ALL" else [machine_sel]

        for sheet_name in target_sheets:
            if sheet_name not in wb.sheetnames:
                buf.write(f"Skipping {sheet_name} — sheet not found.\n")
                continue
            ws = wb[sheet_name]
            for d in all_dates:
                row_num = bot_main.find_date_row(ws, d)
                if row_num == 0:
                    row_num = bot_main.add_date_row(ws, d, "B", "DS")
                    filled += 1
                # Ensure all numeric cells are 0 where blank
                for start_col in bot_main.CREW_START_COL.values():
                    for offset in list(range(bot_main.COL_FAULTS_START, bot_main.COL_FAULTS_END + 1)) + [bot_main.COL_TONS, bot_main.COL_M2255]:
                        cell = ws.cell(row_num, start_col + offset)
                        if cell.value is None or cell.value == "":
                            cell.value = 0

        wb.save(output_path)
        machine_label = machine_sel if machine_sel != "ALL" else "all machines"
        buf.write(f"Filled {filled} new date row(s) with zeros on {machine_label}.\n")
        buf.write(f"All cells in date range {date_from} → {date_to} now have at least 0.\n")
        buf.write(f"✓ Saved: {output_path}\n")

    except ValueError as e:
        ok = False
        buf.write(f"ERROR: Invalid date format — use YYYY-MM-DD. ({e})\n")
    except FileNotFoundError as e:
        ok = False
        buf.write(f"ERROR: {e}\n")
    except Exception as e:
        ok = False
        buf.write(f"UNEXPECTED ERROR: {e}\n")
        import traceback
        buf.write(traceback.format_exc())

    return jsonify({"ok": ok, "log": buf.getvalue()})


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import threading
    import webbrowser

    port = 5050
    url  = f"http://localhost:{port}"

    def open_browser():
        webbrowser.open(url)

    threading.Timer(1.2, open_browser).start()

    print(f"╔══════════════════════════════════════════╗")
    print(f"║  Data Bot  —  {url}        ║")
    print(f"║  Press Ctrl+C to stop the server.        ║")
    print(f"╚══════════════════════════════════════════╝")

    app.run(host="0.0.0.0", port=port, debug=False)
