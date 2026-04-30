"""
generate_report.py
------------------
Reads 'Faults vs Tons 2026.xlsx', computes per-rotation (7-shift block) summaries
per crew for every machine sheet, then writes a new report file that exactly mirrors
the summary-area layout and 14-chart structure already built in CM02/CM03/CM05.

Column layout (matching CM02):
  Cols 71-74  (BS-BV): Total tons per rotation, B/D/C/A crews
  Cols 83-86  (CE-CH): Tons/Fault per rotation
  Cols 95-98  (CQ-CT): Accumulated tons (running SUM)
  Cols 107-110(DC-DF): Accumulated Tons/Fault (running SUM)
  Cols 119-120(DO-DP): B Crew tons + running avg
  Cols 129-130(DY-DZ): D Crew tons + running avg
  Cols 139-140(EI-EJ): C Crew tons + running avg
  Cols 149-150(ES-ET): A Crew tons + running avg
  Cols 159-162(FC-FF): All-crew running avg tons
  Cols 171-172(FO-FP): B Crew Tons/Fault + running avg
  Cols 181-182(FY-FZ): D Crew Tons/Fault + running avg
  Cols 191-192(GI-GJ): C Crew Tons/Fault + running avg
  Cols 201-202(GS-GT): A Crew Tons/Fault + running avg
  Cols 211-214(HC-HF): All-crew running avg Tons/Fault
"""

import os
import sys
import openpyxl
from openpyxl import Workbook
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.chart.data_source import StrRef, StrVal, StrData
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── Constants ──────────────────────────────────────────────────────────────
MACHINE_SHEETS   = ["CM02", "CM03", "CM05", "CM07", "CM08", "CM09"]
CREW_ORDER       = ["B", "D", "C", "A"]
CREW_START_COL   = {"B": 2, "D": 19, "C": 36, "A": 53}
COL_SHIFT        = 0
COL_DATE         = 1
COL_FAULTS_START = 2
COL_FAULTS_END   = 13
COL_TOTAL        = 14
COL_TONS         = 15
DATA_START_ROW   = 5
ROTATION_SIZE    = 7

# Summary column positions (matching CM02)
G1 = {"B": 71, "D": 72, "C": 73, "A": 74}      # BS-BV  total tons/rotation
G2 = {"B": 83, "D": 84, "C": 85, "A": 86}      # CE-CH  tons/fault per rotation
G3 = {"B": 95, "D": 96, "C": 97, "A": 98}      # CQ-CT  accumulated tons
G4 = {"B":107, "D":108, "C":109, "A":110}      # DC-DF  accumulated tons/fault
G5 = {"B":119, "D":129, "C":139, "A":149}      # DO/DY/EI/ES individual tons (2-col: value + avg)
G9 = {"B":159, "D":160, "C":161, "A":162}      # FC-FF  all-crew avg tons
G10= {"B":171, "D":181, "C":191, "A":201}      # FO/FY/GI/GS individual T/F (2-col: value + avg)
G14= {"B":211, "D":212, "C":213, "A":214}      # HC-HF  all-crew avg T/F

HDR_ROW   = 5
DAT_START = 6

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
SUBHEAD_FILL= PatternFill("solid", fgColor="2E75B6")
ALT_FILL    = PatternFill("solid", fgColor="D6E4F0")
WF          = Font(color="FFFFFF", bold=True, size=10)
BF          = Font(bold=True)

CREW_LABEL = {"B":"B Crew","D":"D Crew","C":"C Crew","A":"A Crew"}


def _sf(v):
    try:
        return float(v) if v not in (None,""," ") else 0.0
    except (ValueError,TypeError):
        return 0.0


def read_crew_shifts(ws, crew):
    sc = CREW_START_COL[crew]
    out = []
    for rn in range(DATA_START_ROW, ws.max_row+1):
        date_v = ws.cell(rn, sc+COL_DATE).value
        shft_v = ws.cell(rn, sc+COL_SHIFT).value
        if date_v is None or shft_v is None:
            continue
        tons   = _sf(ws.cell(rn, sc+COL_TONS).value)
        faults = _sf(ws.cell(rn, sc+COL_TOTAL).value)
        out.append((tons, faults))
    return out


def compute_rotations(shifts):
    rots = []
    for i in range(0, len(shifts), ROTATION_SIZE):
        block = shifts[i:i+ROTATION_SIZE]
        if not block:
            continue
        tt = sum(t for t,f in block)
        tf_total = sum(f for t,f in block)
        avg_tf = round(tt/tf_total, 4) if tf_total > 0 else 0.0
        rots.append({"tt": tt, "tf": avg_tf})
    return rots


def hcell(ws, row, col, val, fill=None, font=None, align="center"):
    c = ws.cell(row, col, val)
    if fill: c.fill = fill
    if font: c.font = font
    c.alignment = Alignment(horizontal=align)
    return c


def write_summary(ws, all_rots):
    n = max((len(v) for v in all_rots.values()), default=0)
    if n == 0:
        return 0
    cl = get_column_letter

    # Header row
    for crew in CREW_ORDER:
        for gd in [G1,G2,G3,G4,G9,G14]:
            hcell(ws, HDR_ROW, gd[crew], CREW_LABEL[crew], HEADER_FILL, WF)
        for gd in [G5,G10]:
            hcell(ws, HDR_ROW, gd[crew],   CREW_LABEL[crew], HEADER_FILL, WF)
            lbl = "Avr. Tons" if gd is G5 else "Avr. Fault"
            hcell(ws, HDR_ROW, gd[crew]+1, lbl,              HEADER_FILL, WF)

    # Data rows
    for i in range(n):
        row = DAT_START + i

        for crew in CREW_ORDER:
            rots = all_rots.get(crew, [])
            tt = rots[i]["tt"] if i < len(rots) else 0.0
            tf = rots[i]["tf"] if i < len(rots) else 0.0

            # G1 total tons
            ws.cell(row, G1[crew], round(tt,2))
            # G2 tons/fault
            ws.cell(row, G2[crew], round(tf,4))

            # G3 accumulated tons
            g3c = cl(G3[crew])
            if i == 0:
                ws.cell(row, G3[crew], round(tt,2))
            else:
                ws.cell(row, G3[crew], f"=SUM({g3c}{row-1},{cl(G1[crew])}{row})")

            # G4 accumulated T/F
            if i == 0:
                ws.cell(row, G4[crew], round(tf,4))
            else:
                ws.cell(row, G4[crew],
                        f"=SUM({cl(G4[crew])}{row-1},{cl(G2[crew])}{row})")

            # G5 individual + running avg
            b5 = G5[crew]
            if i == 0:
                ws.cell(row, b5, round(tt,2))
            else:
                ws.cell(row, b5, f"={cl(G1[crew])}{row}")
            ws.cell(row, b5+1,
                    f"=AVERAGE({cl(b5)}{DAT_START}:{cl(b5)}{row})")

            # G10 individual + running avg
            b10 = G10[crew]
            if i == 0:
                ws.cell(row, b10, round(tf,4))
            else:
                ws.cell(row, b10, f"={cl(G2[crew])}{row}")
            ws.cell(row, b10+1,
                    f"=AVERAGE({cl(b10)}{DAT_START}:{cl(b10)}{row})")

        # G9 = running avg from G5 avg column
        for crew in CREW_ORDER:
            ws.cell(row, G9[crew], f"={cl(G5[crew]+1)}{row}")

        # G14 = running avg from G10 avg column
        for crew in CREW_ORDER:
            ws.cell(row, G14[crew], f"={cl(G10[crew]+1)}{row}")

    return n


def add_charts(ws, sheet_name, n):
    if n == 0:
        return
    d0 = DAT_START
    d1 = DAT_START + n - 1

    def ref(col):
        return Reference(ws, min_col=col, max_col=col, min_row=d0, max_row=d1)

    def sl(col):
        return SeriesLabel(strRef=StrRef(
            f=f"'{sheet_name}'!${get_column_letter(col)}${HDR_ROW}",
            strCache=StrData(ptCount=1,
                             pt=[StrVal(idx=0,v=get_column_letter(col))])
        ))

    def line4(title, gd, anchor):
        c = LineChart()
        c.title = title; c.style = 10; c.width = 20; c.height = 10
        for crew in CREW_ORDER:
            c.add_data(ref(gd[crew]), titles_from_data=False)
            c.series[-1].title = sl(gd[crew])
        ws.add_chart(c, anchor)

    def bar1(col, anchor):
        c = BarChart()
        c.type="col"; c.style=10; c.width=13; c.height=9
        c.add_data(ref(col), titles_from_data=False)
        c.series[-1].title = sl(col)
        ws.add_chart(c, anchor)

    # 14 charts matching CM02 order
    line4("Tons per Rotation",              G1,  "BX2")
    line4("Faults/Tons per Rotation",       G2,  "BX20")
    line4("Tons per Shift Accumulated",     G3,  "BX38")
    line4("Faults/Tons Accumulated",        G4,  "BX56")
    bar1(G5["B"],                                "BX74")
    bar1(G5["D"],                                "BX92")
    bar1(G5["C"],                                "BX110")
    bar1(G5["A"],                                "BX128")
    line4("Tons per Shift Average",         G9,  "BX146")
    bar1(G10["A"],                               "BX164")
    line4("Faults/Tons per Shift Average",  G14, "BX182")
    bar1(G10["B"],                               "BX200")
    bar1(G10["C"],                               "BX218")
    bar1(G10["D"],                               "BX236")


def write_combined(report_wb, summaries):
    ws = report_wb.create_sheet("Combined", 0)
    n = max(
        (len(rots) for m in summaries.values() for rots in m.values()),
        default=0
    )
    if n == 0:
        return

    ws.merge_cells("A1:N1")
    c = ws["A1"]
    c.value = "Site Combined — All Machines per Rotation"
    c.font  = Font(size=13, bold=True, color="FFFFFF")
    c.fill  = HEADER_FILL
    c.alignment = Alignment(horizontal="center")

    HDR = 2; DS = 3
    hcell(ws, HDR, 1, "Rotation", SUBHEAD_FILL, WF)
    col = 2
    mcols = {}
    for m in MACHINE_SHEETS:
        mcols[m] = col
        hcell(ws, HDR, col,   f"{m} Tons",    SUBHEAD_FILL, WF)
        hcell(ws, HDR, col+1, f"{m} T/Fault", SUBHEAD_FILL, WF)
        col += 2
    site_t  = col;     hcell(ws, HDR, site_t,   "Site Tons",    HEADER_FILL, WF)
    site_tf = col + 1; hcell(ws, HDR, site_tf,  "Site T/Fault", HEADER_FILL, WF)

    for i in range(n):
        row  = DS + i
        fill = ALT_FILL if i % 2 == 0 else None
        ws.cell(row, 1, f"R{i+1}").font = BF
        if fill: ws.cell(row, 1).fill = fill

        st = 0.0; sf = 0.0
        for m in MACHINE_SHEETS:
            mt = mf = 0.0
            for crew in CREW_ORDER:
                rots = summaries.get(m,{}).get(crew,[])
                if i < len(rots):
                    mt += rots[i]["tt"]
                    mf += rots[i]["tt"] / rots[i]["tf"] if rots[i]["tf"] > 0 else 0.0
            mtf = round(mt/mf, 2) if mf > 0 else 0.0
            c1 = ws.cell(row, mcols[m],   round(mt,1))
            c2 = ws.cell(row, mcols[m]+1, mtf)
            if fill: c1.fill = fill; c2.fill = fill
            st += mt; sf += mf

        stf = round(st/sf, 2) if sf > 0 else 0.0
        c3 = ws.cell(row, site_t,  round(st,1)); c3.font = BF
        c4 = ws.cell(row, site_tf, stf);         c4.font = BF
        if fill: c3.fill = fill; c4.fill = fill

    cat = Reference(ws, min_col=1, max_col=1, min_row=DS, max_row=DS+n-1)
    last = site_tf + 3

    def cline(title, col, anchor):
        c = LineChart(); c.title=title; c.style=10; c.width=25; c.height=14
        c.add_data(Reference(ws,min_col=col,max_col=col,min_row=HDR,max_row=DS+n-1),
                   titles_from_data=True)
        c.set_categories(cat)
        ws.add_chart(c, anchor)

    cline("Site Total Tonnes per Rotation",      site_t,  f"{get_column_letter(last)}2")
    cline("Site Tonnes per Fault per Rotation",  site_tf, f"{get_column_letter(last)}26")

    cb = LineChart(); cb.title="Tons per Miner per Rotation"
    cb.style=10; cb.width=25; cb.height=14
    for m in MACHINE_SHEETS:
        mc = mcols[m]
        cb.add_data(Reference(ws,min_col=mc,max_col=mc,min_row=HDR,max_row=DS+n-1),
                    titles_from_data=True)
    cb.set_categories(cat)
    ws.add_chart(cb, f"{get_column_letter(last+17)}2")

    cd = LineChart(); cd.title="Tonnes per Fault per Miner"
    cd.style=10; cd.width=25; cd.height=14
    for m in MACHINE_SHEETS:
        mc = mcols[m]+1
        cd.add_data(Reference(ws,min_col=mc,max_col=mc,min_row=HDR,max_row=DS+n-1),
                    titles_from_data=True)
    cd.set_categories(cat)
    ws.add_chart(cd, f"{get_column_letter(last+17)}26")


def generate_report(source_path: str, output_path: str) -> str:
    print(f"Loading: {source_path}")
    wb_src = openpyxl.load_workbook(source_path, data_only=True)

    report_wb = Workbook()
    report_wb.remove(report_wb.active)
    summaries = {}

    log = []
    for machine in MACHINE_SHEETS:
        if machine not in wb_src.sheetnames:
            log.append(f"  {machine}: sheet not found, skipped")
            continue
        print(f"  Processing {machine}...")
        ws_src = wb_src[machine]
        ws_rep = report_wb.create_sheet(machine)

        # Copy header + raw data rows
        for rn in range(1, ws_src.max_row + 1):
            for cn in range(1, 70):
                v = ws_src.cell(rn, cn).value
                if v is not None:
                    ws_rep.cell(rn, cn, v)

        all_rots = {}
        for crew in CREW_ORDER:
            shifts = read_crew_shifts(ws_src, crew)
            all_rots[crew] = compute_rotations(shifts)
        summaries[machine] = all_rots

        n = write_summary(ws_rep, all_rots)
        add_charts(ws_rep, machine, n)

        total_t = sum(r["tt"] for rots in all_rots.values() for r in rots)
        msg = f"  {machine}: {total_t:,.1f} tons, {n} rotation rows"
        print(msg); log.append(msg)

    write_combined(report_wb, summaries)
    report_wb.save(output_path)
    print(f"\nSaved → {output_path}")
    log.append(f"\n✓ Report saved: {output_path}")
    return "\n".join(log)


if __name__ == "__main__":
    src = sys.argv[1] if len(sys.argv) > 1 else "Faults vs Tons 2026.xlsx"
    out = sys.argv[2] if len(sys.argv) > 2 else "Faults vs Tons Report.xlsx"
    generate_report(src, out)
