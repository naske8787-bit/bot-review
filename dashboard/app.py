import os
import sys
import csv
import io
import secrets
from datetime import datetime, timedelta
from flask import Flask, send_from_directory, request, jsonify, abort
from db import init_db, get_conn

app = Flask(__name__, static_folder='static')

MACHINES      = ["CM02", "CM03", "CM05", "CM07", "CM08", "CM09"]
CREWS         = ["B", "D", "C", "A"]
ROTATION_SIZE = 7
ROTATION_GAP_DAYS = 2

# ── Read-only share token ─────────────────────────────────────────────────
# Set DASHBOARD_VIEW_TOKEN env var to a fixed value, otherwise one is
# generated fresh each restart (recipients will need a new link after restart).
_VIEW_TOKEN = os.environ.get("DASHBOARD_VIEW_TOKEN") or secrets.token_urlsafe(24)
print(f"  Read-only share URL path: /view/{_VIEW_TOKEN}")

def _is_readonly_token(token):
    return secrets.compare_digest(token, _VIEW_TOKEN)

def _block_writes():
    """Call at the start of any write endpoint to reject read-only access."""
    # Read-only viewers come from /view/<token> which sets a cookie; direct
    # API calls without the write header from the owner UI are also blocked.
    if request.headers.get("X-Write-Auth") != _VIEW_TOKEN:
        abort(403, description="Read-only mode: modifications are not allowed.")


def _split_rotations(rows, date_key='date'):
    """Split a list of shift dicts into rotation blocks based on date gaps."""
    if not rows:
        return []
    blocks, current = [], [rows[0]]
    for s in rows[1:]:
        prev = datetime.strptime(current[-1][date_key], '%Y-%m-%d')
        curr = datetime.strptime(s[date_key], '%Y-%m-%d')
        if (curr - prev).days > ROTATION_GAP_DAYS:
            blocks.append(current)
            current = []
        current.append(s)
    blocks.append(current)
    return blocks


# Fixed fault-code columns shown in every shift table
FAULT_CODES = [
    "M1441", "M1443", "M1447", "M1448", "M1450",
    "M1452", "M1454", "M1456", "M1458", "M1461", "M1463",
]


_HERE = os.path.dirname(os.path.abspath(__file__))
_REPO_ROOT = os.path.dirname(_HERE)
_TRADING_EQUITY_LOG = os.path.join(_REPO_ROOT, 'trading_bot', 'logs', 'equity_log.csv')
_CRYPTO_TRADE_LOG = os.path.join(_REPO_ROOT, 'crypto_bot', 'logs', 'trade_log.csv')


def _parse_iso_timestamp(value):
    if not value:
        return None
    text = str(value).strip()
    if not text:
        return None
    try:
        return datetime.fromisoformat(text.replace('Z', '+00:00'))
    except ValueError:
        return None


def _read_trading_equity_series(path):
    out = []
    if not os.path.exists(path):
        return out
    try:
        with open(path, 'r', encoding='utf-8', newline='') as f:
            reader = csv.DictReader(f)
            for row in reader:
                ts_raw = row.get('timestamp')
                ts = _parse_iso_timestamp(ts_raw)
                if ts is None:
                    continue
                try:
                    portfolio_value = float(row.get('portfolio_value', 0.0) or 0.0)
                    cash_balance = float(row.get('cash_balance', 0.0) or 0.0)
                    buying_power = float(row.get('buying_power', cash_balance) or cash_balance)
                except (TypeError, ValueError):
                    continue
                out.append({
                    't': ts.isoformat(),
                    'portfolio_value': round(portfolio_value, 2),
                    'cash_balance': round(cash_balance, 2),
                    'buying_power': round(buying_power, 2),
                })
    except Exception:
        return []
    return sorted(out, key=lambda x: x['t'])


def _read_crypto_realized_pnl_series(path):
    out = []
    if not os.path.exists(path):
        return out
    running = 0.0
    rows = []
    try:
        with open(path, 'r', encoding='utf-8', newline='') as f:
            reader = csv.DictReader(f)
            for row in reader:
                ts = _parse_iso_timestamp(row.get('exit_time'))
                if ts is None:
                    continue
                try:
                    pnl = float(row.get('pnl', 0.0) or 0.0)
                except (TypeError, ValueError):
                    pnl = 0.0
                rows.append((ts, pnl))
    except Exception:
        return []

    for ts, pnl in sorted(rows, key=lambda x: x[0]):
        running += pnl
        out.append({'t': ts.isoformat(), 'cum_realized_pnl': round(running, 2)})
    return out

# ── Static ────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    """Owner view — full edit access."""
    with open(os.path.join(_HERE, 'static', 'index.html'), 'r') as f:
        html = f.read()
    # Inject the write token so the owner's browser can call write endpoints
    html = html.replace('</head>', f'<script>window._WRITE_TOKEN="{_VIEW_TOKEN}";window.READONLY=false;</script>\n</head>', 1)
    return html


@app.route('/view/<token>')
def view_readonly(token):
    """Read-only share link for email recipients."""
    if not _is_readonly_token(token):
        abort(403, description="Invalid or expired share link.")
    with open(os.path.join(_HERE, 'static', 'index.html'), 'r') as f:
        html = f.read()
    html = html.replace('</head>', '<script>window._WRITE_TOKEN=null;window.READONLY=true;</script>\n</head>', 1)
    return html


@app.route('/api/share-link')
def share_link():
    """Return the read-only share URL (owner only, requires write token)."""
    if request.headers.get("X-Write-Auth") != _VIEW_TOKEN:
        abort(403)
    host = request.host_url.rstrip('/')
    return jsonify({'url': f"{host}/view/{_VIEW_TOKEN}"})


# ── Fault codes known for a machine (autocomplete) ───────────────────────

@app.route('/api/fault-codes')
def fault_codes():
    machine = request.args.get('machine', '')
    with get_conn() as c:
        if machine:
            rows = c.execute("""
                SELECT DISTINCT sf.fault_code
                FROM shift_faults sf
                JOIN shifts s ON s.id = sf.shift_id
                WHERE s.machine = ?
                ORDER BY sf.fault_code
            """, (machine,)).fetchall()
        else:
            rows = c.execute(
                "SELECT DISTINCT fault_code FROM shift_faults ORDER BY fault_code"
            ).fetchall()
    return jsonify([r['fault_code'] for r in rows])


# ── All shifts for a machine+crew, grouped into 7-shift rotations ────────

@app.route('/api/shifts')
def get_shifts():
    machine = request.args.get('machine', '')
    crew    = request.args.get('crew', '')
    if not machine or not crew:
        return jsonify({'error': 'machine and crew required'}), 400

    with get_conn() as c:
        shifts = c.execute("""
            SELECT id, date, shift, tons FROM shifts
            WHERE machine = ? AND crew = ?
            ORDER BY date, CASE shift WHEN 'DS' THEN 1 ELSE 2 END
        """, (machine, crew)).fetchall()

        sid_list = [s['id'] for s in shifts]
        fault_map = {}
        if sid_list:
            ph = ','.join('?' * len(sid_list))
            for f in c.execute(
                f"SELECT shift_id, fault_code, count FROM shift_faults "
                f"WHERE shift_id IN ({ph})",
                sid_list
            ).fetchall():
                fault_map.setdefault(f['shift_id'], {})[f['fault_code']] = f['count']

    result = []
    for s in shifts:
        f = fault_map.get(s['id'], {})
        result.append({
            'id':           s['id'],
            'date':         s['date'],
            'shift':        s['shift'],
            'tons':         s['tons'],
            'faults':       f,
            'total_faults': sum(f.values()),
        })

    rotations = []
    for idx, block in enumerate(_split_rotations(result), start=1):
        tt = sum(s['tons'] for s in block)
        tf = sum(s['total_faults'] for s in block)
        ft = {}
        for s in block:
            for code, cnt in s['faults'].items():
                ft[code] = ft.get(code, 0) + cnt
        n = len(block)
        rotations.append({
            'rotation':     idx,
            'shifts':       block,
            'total_tons':   round(tt, 2),
            'avg_tons':     round(tt / n, 2) if n else 0,
            'total_faults': tf,
            'avg_faults':   round(tf / n, 2) if n else 0,
            'fault_totals': ft,
            'tons_per_fault': round(tt / tf, 2) if tf > 0 else 0,
            'complete':     n == ROTATION_SIZE,
        })

    return jsonify({
        'rotations':    rotations,
        'fault_codes':  FAULT_CODES,
        'total_shifts': len(result),
    })


# ── All-crews summary for a single machine (per rotation) ────────────────

@app.route('/api/machine-summary')
def machine_summary():
    machine = request.args.get('machine', '')
    if not machine:
        return jsonify({'error': 'machine required'}), 400

    with get_conn() as c:
        result = {}
        for crew in CREWS:
            rows = c.execute("""
                SELECT s.id, s.date, s.tons, COALESCE(SUM(sf.count), 0) AS tf
                FROM shifts s
                LEFT JOIN shift_faults sf ON sf.shift_id = s.id
                WHERE s.machine = ? AND s.crew = ?
                GROUP BY s.id
                ORDER BY s.date, CASE s.shift WHEN 'DS' THEN 1 ELSE 2 END
            """, (machine, crew)).fetchall()

            rots = []
            for idx, blk in enumerate(_split_rotations(rows), start=1):
                tt = sum(r['tons'] for r in blk)
                tf = sum(r['tf'] for r in blk)
                rots.append({
                    'rotation': idx,
                    'total_tons': round(tt, 2),
                    'total_faults': tf,
                    'avg_tons': round(tt / len(blk), 2) if blk else 0,
                    'tons_per_fault': round(tt / tf, 2) if tf > 0 else 0,
                    'n': len(blk),
                })
            result[crew] = rots

    return jsonify(result)


# ── Add or update a shift ────────────────────────────────────────────────

@app.route('/api/shift', methods=['POST'])
def upsert_shift():
    _block_writes()
    d       = request.json or {}
    machine = str(d.get('machine', '')).strip()
    crew    = str(d.get('crew', '')).strip()
    date    = str(d.get('date', '')).strip()
    shift   = str(d.get('shift', '')).strip()
    tons    = float(d.get('tons') or 0)
    faults  = d.get('faults', {})

    if not all([machine, crew, date, shift]):
        return jsonify({'error': 'machine, crew, date, shift required'}), 400
    if machine not in MACHINES:
        return jsonify({'error': 'invalid machine'}), 400
    if crew not in CREWS:
        return jsonify({'error': 'invalid crew'}), 400
    if shift not in ('DS', 'NS'):
        return jsonify({'error': 'shift must be DS or NS'}), 400

    with get_conn() as c:
        c.execute("""
            INSERT INTO shifts (machine, crew, date, shift, tons)
            VALUES (?, ?, ?, ?, ?)
            ON CONFLICT(machine, crew, date, shift)
            DO UPDATE SET tons = excluded.tons
        """, (machine, crew, date, shift, tons))

        sid = c.execute("""
            SELECT id FROM shifts
            WHERE machine=? AND crew=? AND date=? AND shift=?
        """, (machine, crew, date, shift)).fetchone()['id']

        # Replace faults
        c.execute("DELETE FROM shift_faults WHERE shift_id = ?", (sid,))
        for code, count in faults.items():
            code  = str(code).strip().upper()
            try:
                count = int(count)
            except (ValueError, TypeError):
                count = 0
            if code and count > 0:
                c.execute("""
                    INSERT INTO shift_faults (shift_id, fault_code, count)
                    VALUES (?, ?, ?)
                """, (sid, code, count))

    return jsonify({'success': True, 'id': sid})


# ── Delete a shift ───────────────────────────────────────────────────────

@app.route('/api/shift/<int:sid>', methods=['DELETE'])
def delete_shift(sid):
    _block_writes()
    with get_conn() as c:
        c.execute("DELETE FROM shifts WHERE id = ?", (sid,))
    return jsonify({'success': True})


@app.route('/api/shifts', methods=['DELETE'])
def clear_crew_shifts():
    """Delete all shifts (and their faults) for a given machine + crew."""
    _block_writes()
    machine = request.args.get('machine', '').strip()
    crew    = request.args.get('crew', '').strip()
    if not machine or not crew:
        return jsonify({'error': 'machine and crew query params required'}), 400
    if machine not in MACHINES:
        return jsonify({'error': 'invalid machine'}), 400
    if crew not in CREWS:
        return jsonify({'error': 'invalid crew'}), 400
    with get_conn() as c:
        ids = [r['id'] for r in c.execute(
            "SELECT id FROM shifts WHERE machine=? AND crew=?", (machine, crew)
        ).fetchall()]
        deleted_faults = 0
        if ids:
            placeholders = ','.join('?' * len(ids))
            deleted_faults = c.execute(
                f"DELETE FROM shift_faults WHERE shift_id IN ({placeholders})", ids
            ).rowcount
            c.execute(
                f"DELETE FROM shifts WHERE id IN ({placeholders})", ids
            )
            c.execute(
                "DELETE FROM raw_fault_events WHERE machine=? AND crew=?", (machine, crew)
            )
    return jsonify({'success': True, 'deleted_shifts': len(ids), 'deleted_faults': deleted_faults})


# ── Combined view: all machines × all crews ──────────────────────────────

@app.route('/api/combined')
def combined():
    result = {}
    with get_conn() as c:
        for machine in MACHINES:
            m = {}
            for crew in CREWS:
                rows = c.execute("""
                    SELECT s.date, s.tons, COALESCE(SUM(sf.count), 0) AS tf
                    FROM shifts s
                    LEFT JOIN shift_faults sf ON sf.shift_id = s.id
                    WHERE s.machine = ? AND s.crew = ?
                    GROUP BY s.id
                    ORDER BY s.date, CASE s.shift WHEN 'DS' THEN 1 ELSE 2 END
                """, (machine, crew)).fetchall()

                rots = []
                for blk in _split_rotations(rows):
                    tt = sum(r['tons'] for r in blk)
                    tf = sum(r['tf']   for r in blk)
                    rots.append({'tt': round(tt, 2), 'tf': tf, 'n': len(blk)})
                m[crew] = rots
            result[machine] = m

    return jsonify(result)


@app.route('/api/investment-performance')
def investment_performance():
    """Return investment progress series from bot logs for dashboard charting."""
    try:
        points = int(request.args.get('points', '240'))
    except ValueError:
        points = 240
    points = max(30, min(points, 1000))

    equity_rows = _read_trading_equity_series(_TRADING_EQUITY_LOG)
    crypto_rows = _read_crypto_realized_pnl_series(_CRYPTO_TRADE_LOG)

    if len(equity_rows) > points:
        equity_rows = equity_rows[-points:]
    if len(crypto_rows) > points:
        crypto_rows = crypto_rows[-points:]

    latest_portfolio = equity_rows[-1]['portfolio_value'] if equity_rows else 0.0
    latest_cash = equity_rows[-1]['cash_balance'] if equity_rows else 0.0
    latest_buying_power = equity_rows[-1]['buying_power'] if equity_rows else 0.0
    latest_crypto_pnl = crypto_rows[-1]['cum_realized_pnl'] if crypto_rows else 0.0
    start_portfolio = equity_rows[0]['portfolio_value'] if equity_rows else 0.0
    net_change = latest_portfolio - start_portfolio if equity_rows else 0.0
    pct_change = (net_change / start_portfolio) if start_portfolio > 0 else 0.0

    return jsonify({
        'portfolio': [
            {'t': row['t'], 'v': row['portfolio_value']} for row in equity_rows
        ],
        'cash': [
            {'t': row['t'], 'v': row['cash_balance']} for row in equity_rows
        ],
        'buying_power': [
            {'t': row['t'], 'v': row['buying_power']} for row in equity_rows
        ],
        'crypto_pnl': [
            {'t': row['t'], 'v': row['cum_realized_pnl']} for row in crypto_rows
        ],
        'latest': {
            'portfolio_value': round(latest_portfolio, 2),
            'cash_balance': round(latest_cash, 2),
            'buying_power': round(latest_buying_power, 2),
            'crypto_cum_realized_pnl': round(latest_crypto_pnl, 2),
            'window_net_change': round(net_change, 2),
            'window_pct_change': round(pct_change * 100, 2),
        },
    })


@app.route('/api/crypto-influencer')
def crypto_influencer():
    """
    Return the latest influencer manipulation signals from the crypto bot's
    research snapshot, plus recent trade log entries tagged as pump-ride.
    Reads the cached research file written by data_fetcher — no live search
    call is made from the dashboard.
    """
    import json as _json

    _RESEARCH_SNAPSHOT = os.path.join(
        _REPO_ROOT, 'tech_research_bot', 'output', 'latest_research.json'
    )
    _CRYPTO_BOT_DIR = os.path.join(_REPO_ROOT, 'crypto_bot')

    # ── influencer snapshot from crypto_bot models cache ──────────────────
    influencer_cache_path = os.path.join(
        _CRYPTO_BOT_DIR, 'models', 'influencer_cache.json'
    )
    # Also try the analysis output written by the bot loop
    bot_analysis_path = os.path.join(
        _CRYPTO_BOT_DIR, 'logs', 'influencer_analysis.json'
    )

    influencer_data = {"by_symbol": {}, "global": {}}
    for candidate in (bot_analysis_path, influencer_cache_path):
        if os.path.exists(candidate):
            try:
                with open(candidate, 'r', encoding='utf-8') as f:
                    influencer_data = _json.load(f)
                break
            except Exception:
                pass

    # ── pump-ride trades from trade log ───────────────────────────────────
    pump_trades = []
    if os.path.exists(_CRYPTO_TRADE_LOG):
        try:
            import csv as _csv
            with open(_CRYPTO_TRADE_LOG, 'r', encoding='utf-8', newline='') as f:
                for row in _csv.DictReader(f):
                    if str(row.get('pump_mode', '') or '').lower() in ('true', '1', 'yes'):
                        pump_trades.append({
                            'symbol': row.get('symbol', ''),
                            'entry_time': row.get('entry_time', ''),
                            'exit_time': row.get('exit_time', ''),
                            'pnl': float(row.get('pnl', 0) or 0),
                            'influencer_actors': row.get('influencer_actors', ''),
                        })
        except Exception:
            pass

    # ── tech research highlights ──────────────────────────────────────────
    research_headlines = []
    if os.path.exists(_RESEARCH_SNAPSHOT):
        try:
            with open(_RESEARCH_SNAPSHOT, 'r', encoding='utf-8') as f:
                snap = _json.load(f)
            research_headlines = [
                {'title': item.get('title', ''), 'probability': item.get('probability', 0.0)}
                for item in (snap if isinstance(snap, list) else snap.get('items', []))[:10]
            ]
        except Exception:
            pass

    return jsonify({
        'influencer': influencer_data,
        'pump_trades': pump_trades[-20:],
        'research_headlines': research_headlines,
        'as_of': datetime.utcnow().isoformat() + 'Z',
    })



# ── Shift allocation from timestamp ─────────────────────────────────────
# DS = Day Shift  06:00 (inclusive) → 18:00 (exclusive)
# NS = Night Shift 18:00 → 06:00 next day
#
# For a fault at 2026-01-07 14:32  → date=2026-01-07, shift=DS
# For a fault at 2026-01-07 22:15  → date=2026-01-07, shift=NS
# For a fault at 2026-01-08 03:10  → date=2026-01-07, shift=NS  (pre-06:00 belongs to prev day NS)

def allocate_shift(dt: datetime):
    """Return (date_str, shift) for a given datetime."""
    h = dt.hour
    if 6 <= h < 18:
        return dt.strftime('%Y-%m-%d'), 'DS'
    elif h >= 18:
        return dt.strftime('%Y-%m-%d'), 'NS'
    else:
        # 00:00 – 05:59 → belongs to previous day's NS
        prev = dt - timedelta(days=1)
        return prev.strftime('%Y-%m-%d'), 'NS'


def _parse_ts(s: str):
    """Try several common timestamp formats."""
    for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M',
                '%d/%m/%Y %H:%M:%S', '%d/%m/%Y %H:%M',
                '%Y/%m/%d %H:%M:%S', '%Y/%m/%d %H:%M'):
        try:
            return datetime.strptime(s.strip(), fmt)
        except ValueError:
            continue
    raise ValueError(f"Unrecognised timestamp format: '{s}'")


# ── Import fault log (timestamped CSV / text) ────────────────────────────
#
# Accepted POST body (JSON):
#   machine   : "CM02"
#   crew      : "D"
#   text      : raw CSV / TSV / space-separated lines
#               Each line:  <timestamp>  <fault_code>
#               e.g.  "2026-01-07 14:32, M2255"
#               or    "07/01/2026 14:32:00\tCO101"
#
# The endpoint auto-detects DS vs NS from the timestamp, aggregates
# fault counts per (date, shift, fault_code) and upserts into shift_faults.

@app.route('/api/import-faults', methods=['POST'])
def import_faults():
    _block_writes()
    d            = request.json or {}
    machine      = str(d.get('machine',      '')).strip()
    crew         = str(d.get('crew',         '')).strip()
    text         = str(d.get('text',         '')).strip()
    shift_filter = str(d.get('shift_filter', 'BOTH')).strip().upper()
    if shift_filter not in ('DS', 'NS', 'BOTH'):
        shift_filter = 'BOTH'

    if not all([machine, crew, text]):
        return jsonify({'error': 'machine, crew and text required'}), 400
    if machine not in MACHINES:
        return jsonify({'error': 'invalid machine'}), 400
    if crew not in CREWS:
        return jsonify({'error': 'invalid crew'}), 400

    # Parse lines ──────────────────────────────────────────────────────
    import re
    errors   = []
    buckets  = {}   # (date_str, shift, fault_code) → count

    lines = text.splitlines()

    # ── Detect "Alarm & Event List" multi-column CSV format ──────────
    # Header contains AE_TIMESTAMP and EVENT_CODE columns.
    # We only count STATE=1 rows (alarm raised), skipping STATE=0 (cleared).
    ae_ts_col = ae_code_col = ae_state_col = None
    structured_csv = False

    if lines:
        header_raw = lines[0].strip()
        # strip surrounding quotes from each header field
        header_fields = [f.strip().strip('"').upper()
                         for f in re.split(r',|\t', header_raw)]
        if 'AE_TIMESTAMP' in header_fields and 'EVENT_CODE' in header_fields:
            structured_csv  = True
            ae_ts_col       = header_fields.index('AE_TIMESTAMP')
            ae_code_col     = header_fields.index('EVENT_CODE')
            ae_state_col    = header_fields.index('STATE') if 'STATE' in header_fields else None
            ae_date_col     = header_fields.index('SHIFT_DATE')  if 'SHIFT_DATE'  in header_fields else None
            ae_shift_col    = header_fields.index('SHIFTNAME')   if 'SHIFTNAME'   in header_fields else None

    if structured_csv:
        # Parse multi-column CSV; skip header row
        for lineno, raw in enumerate(lines[1:], 2):
            line = raw.strip()
            if not line:
                continue
            # CSV-aware split (handles quoted fields)
            fields = [f.strip().strip('"') for f in re.split(r',(?=(?:[^"]*"[^"]*")*[^"]*$)', line)]
            if len(fields) <= max(ae_ts_col, ae_code_col):
                errors.append(f"Line {lineno}: too few columns")
                continue
            # Skip alarm-clear rows (STATE=0) — only count rising edges
            if ae_state_col is not None:
                state_val = fields[ae_state_col].strip()
                if state_val == '0':
                    continue
            ts_raw   = fields[ae_ts_col].strip()
            code_raw = fields[ae_code_col].strip().upper()
            if not code_raw:
                errors.append(f"Line {lineno}: empty fault code")
                continue
            # Use SHIFT_DATE + SHIFTNAME directly if available (avoids boundary guessing)
            if ae_date_col is not None and ae_shift_col is not None:
                date_str  = fields[ae_date_col].strip()
                shift_raw = fields[ae_shift_col].strip().upper()
                shift     = 'DS' if shift_raw.startswith('D') else 'NS'
            else:
                try:
                    dt = _parse_ts(ts_raw)
                except ValueError as e:
                    errors.append(str(e))
                    continue
                date_str, shift = allocate_shift(dt)
            # Apply shift filter
            if shift_filter != 'BOTH' and shift != shift_filter:
                continue
            key = (date_str, shift, code_raw)
            buckets[key] = buckets.get(key, 0) + 1
    else:
        # ── Original simple format: <timestamp>, <fault_code> ────────
        for lineno, raw in enumerate(lines, 1):
            line = raw.strip()
            if not line or line.startswith('#'):
                continue
            parts = re.split(r',|\t|  +', line, maxsplit=1)
            if len(parts) < 2:
                errors.append(f"Line {lineno}: cannot split timestamp from fault code — '{line}'")
                continue
            ts_raw, code_raw = parts[0].strip(), parts[1].strip().upper()
            if not code_raw:
                errors.append(f"Line {lineno}: empty fault code")
                continue
            try:
                dt = _parse_ts(ts_raw)
            except ValueError as e:
                errors.append(str(e))
                continue
            date_str, shift = allocate_shift(dt)
            # Apply shift filter
            if shift_filter != 'BOTH' and shift != shift_filter:
                continue
            key = (date_str, shift, code_raw)
            buckets[key] = buckets.get(key, 0) + 1

    if not buckets and errors:
        return jsonify({'error': 'No valid lines parsed', 'details': errors}), 400

    # Write to DB ──────────────────────────────────────────────────────
    saved_shifts = set()
    with get_conn() as c:
        skipped_dates = set()
        for (date_str, shift, fault_code), count in buckets.items():
            # Only add faults to shifts that already exist in the DB.
            # Never auto-create shift rows — user must add shifts via Add Shift first.
            row = c.execute("""
                SELECT id FROM shifts
                WHERE machine=? AND crew=? AND date=? AND shift=?
            """, (machine, crew, date_str, shift)).fetchone()

            if row is None:
                skipped_dates.add((date_str, shift))
                continue

            c.execute("""
                INSERT INTO shift_faults (shift_id, fault_code, count)
                VALUES (?, ?, ?)
                ON CONFLICT(shift_id, fault_code)
                DO UPDATE SET count = count + excluded.count
            """, (row['id'], fault_code, count))

            saved_shifts.add((date_str, shift))

    for date_str, shift in sorted(skipped_dates):
        errors.append(f"No {shift} shift found for {date_str} — add it via Add Shift first")

    shifts_summary = sorted(
        [{'date': d, 'shift': s} for d, s in saved_shifts],
        key=lambda x: (x['date'], x['shift'])
    )

    return jsonify({
        'success':        True,
        'faults_parsed':  sum(v for (d, s, _), v in buckets.items() if (d, s) in saved_shifts),
        'unique_events':  len([k for k in buckets if (k[0], k[1]) in saved_shifts]),
        'shifts_updated': shifts_summary,
        'warnings':       errors,
    })


# ── Import from Excel ────────────────────────────────────────────────────

# Column offsets (0-indexed) for each crew block within a sheet row.
# Each block is 17 columns wide: Shift, Date, 12×fault, Total, TonsCut, M2255
_CREW_BLOCKS = [('B', 1), ('D', 18), ('C', 35), ('A', 52)]

# (within-block offset, fault_code) — M1445 is in the sheet but not tracked
_BLOCK_FAULT_COLS = [
    (2,  'M1443'), (3,  'M1447'), (4,  'M1450'), (5,  'M1454'),
    (6,  'M1458'), (7,  'M1463'), (8,  'M1441'),
    (10, 'M1448'), (11, 'M1452'), (12, 'M1456'), (13, 'M1461'),
    (16, 'M2255'),
]

_EXCEL_PATH = os.path.join(
    os.path.dirname(__file__), '..', 'data_bot', 'Faults vs Tons 2026.xlsx'
)


@app.route('/api/import-excel', methods=['POST'])
def import_excel():
    """Read the pre-placed Excel file and upsert all shift + fault data."""
    _block_writes()
    try:
        import openpyxl
    except ImportError:
        return jsonify({'error': 'openpyxl not installed on server'}), 500

    excel_path = os.path.normpath(_EXCEL_PATH)
    if not os.path.exists(excel_path):
        return jsonify({'error': f'Excel file not found: {excel_path}'}), 404

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
    except Exception as e:
        return jsonify({'error': f'Cannot open workbook: {e}'}), 500

    imported = 0
    skipped  = 0

    with get_conn() as conn:
        for machine in MACHINES:
            if machine not in wb.sheetnames:
                skipped += 1
                continue
            ws = wb[machine]
            for row in ws.iter_rows(values_only=True):
                for crew, offset in _CREW_BLOCKS:
                    # Need at least offset+16 columns
                    if len(row) <= offset + 16:
                        continue
                    shift_val = row[offset]
                    date_val  = row[offset + 1]
                    tons_val  = row[offset + 15]

                    if shift_val not in ('DS', 'NS'):
                        continue
                    if not isinstance(date_val, datetime):
                        continue

                    date_str = date_val.strftime('%Y-%m-%d')
                    try:
                        tons = float(str(tons_val).rstrip('.') if tons_val is not None else 0)
                    except (TypeError, ValueError):
                        tons = 0.0

                    conn.execute("""
                        INSERT INTO shifts (machine, crew, date, shift, tons)
                        VALUES (?, ?, ?, ?, ?)
                        ON CONFLICT(machine, crew, date, shift)
                        DO UPDATE SET tons = excluded.tons
                    """, (machine, crew, date_str, shift_val, tons))

                    sid = conn.execute("""
                        SELECT id FROM shifts
                        WHERE machine=? AND crew=? AND date=? AND shift=?
                    """, (machine, crew, date_str, shift_val)).fetchone()['id']

                    # Replace all fault counts for this shift
                    conn.execute(
                        "DELETE FROM shift_faults WHERE shift_id = ?", (sid,)
                    )
                    for col_off, code in _BLOCK_FAULT_COLS:
                        idx = offset + col_off
                        if idx >= len(row):
                            continue
                        try:
                            count = int(row[idx] or 0)
                        except (TypeError, ValueError):
                            count = 0
                        if count > 0:
                            conn.execute(
                                "INSERT INTO shift_faults (shift_id, fault_code, count)"
                                " VALUES (?, ?, ?)",
                                (sid, code, count)
                            )

                    imported += 1

    return jsonify({'success': True, 'imported': imported})


if __name__ == '__main__':
    init_db()
    print("Mining Dashboard → http://localhost:5051")
    app.run(host='0.0.0.0', port=5051, debug=False)
